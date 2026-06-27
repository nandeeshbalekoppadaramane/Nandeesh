"""
UDIN Extractor - Extract UDIN numbers from scanned PDF reports using RapidOCR.

Features:
- OCR-based extraction using RapidOCR (works on scanned/image PDFs)
- Handles multiple UDIN formats (18-digit alphanumeric)
- Outputs results to Excel (Column A = File Name, Column B = UDIN)
- Moves failed extractions to a "failed extraction" folder

Usage:
    python extract_udin.py [--input-dir INPUT_DIR] [--output OUTPUT_FILE]
"""

import os
import re
import sys
import shutil
import argparse
from pathlib import Path

import fitz  # PyMuPDF - for converting PDF pages to images
from rapidocr_onnxruntime import RapidOCR
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ──────────────────────────────────────────────────────────────────────────────
# UDIN REGEX PATTERNS
# ──────────────────────────────────────────────────────────────────────────────
# UDIN is an 18-character alphanumeric string.
# Format: YY (year) + MMMMMM (6-digit membership) + AAAAAAAAAA (10 alphanumeric)
# Example: 26123456ABCDEF1234
#
# Pattern 1: Standard - UDIN printed as a continuous 18-char string
# Pattern 2: Spaced/Formatted - UDIN may have spaces, dashes, or line breaks
#            e.g., "2612 3456 ABCD EF12 34" or "26-123456-ABCDEF1234"

UDIN_PATTERNS = [
    # Pattern 1: Continuous 18-char alphanumeric (most common)
    # Matches exactly 18 alphanumeric characters that look like a UDIN
    re.compile(r'\b(\d{2}\d{6}[A-Za-z0-9]{10})\b'),

    # Pattern 2: With separators (spaces, dashes, dots)
    # e.g., "26 123456 ABCDEF1234" or "26-123456-ABCDEF1234"
    re.compile(r'\b(\d{2}[\s\-\.]*\d{6}[\s\-\.]*[A-Za-z0-9]{10})\b'),

    # Pattern 3: "UDIN:" or "UDIN No" label followed by the number
    # This catches cases where UDIN is labelled explicitly
    re.compile(
        r'U\s*D\s*I\s*N\s*(?:No\.?|Number|:|\s)*[\s:.\-]*'
        r'([A-Za-z0-9\s\-\.]{16,30})',
        re.IGNORECASE
    ),

    # Pattern 4: Broadly match any 18-char alphanumeric block
    # that starts with 2 digits (year) followed by 6 digits (membership)
    re.compile(r'(\d{2}\d{6}[A-Za-z0-9]{10})'),
]


def clean_udin(raw: str) -> str:
    """Remove spaces, dashes, dots and other separators from extracted UDIN."""
    cleaned = re.sub(r'[\s\-\.\,\:\;]', '', raw)
    return cleaned.strip()


def validate_udin(udin: str) -> bool:
    """
    Validate that a cleaned UDIN looks correct:
    - Exactly 18 characters
    - First 2 chars are digits (year: 19-29 range typically)
    - Next 6 chars are digits (membership number)
    - Last 10 chars are alphanumeric
    """
    if len(udin) != 18:
        return False
    if not udin[:2].isdigit():
        return False
    if not udin[2:8].isdigit():
        return False
    if not udin[8:].isalnum():
        return False
    return True


def extract_udin_from_text(text: str) -> str | None:
    """
    Try all UDIN patterns against the OCR text.
    Returns the first valid UDIN found, or None.
    """
    for pattern in UDIN_PATTERNS:
        matches = pattern.findall(text)
        for match in matches:
            cleaned = clean_udin(match)
            if validate_udin(cleaned):
                return cleaned.upper()

    # Fallback: brute-force scan every 18-char window in cleaned text
    stripped = re.sub(r'\s+', '', text)
    for i in range(len(stripped) - 17):
        window = stripped[i:i + 18]
        if validate_udin(window):
            return window.upper()

    return None


def ocr_pdf(pdf_path: str, ocr_engine: RapidOCR) -> str:
    """
    Convert each page of the PDF to an image and run OCR.
    Returns the concatenated text from all pages.
    """
    all_text = []

    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        print(f"  [ERROR] Could not open PDF: {e}")
        return ""

    for page_num in range(len(doc)):
        page = doc[page_num]

        # Render page to image at 300 DPI for good OCR quality
        pix = page.get_pixmap(dpi=300)
        img_bytes = pix.tobytes("png")

        # Run RapidOCR on the image bytes
        result, _ = ocr_engine(img_bytes)

        if result:
            page_text = " ".join([line[1] for line in result])
            all_text.append(page_text)
            print(f"  [Page {page_num + 1}] OCR extracted {len(page_text)} chars")
        else:
            print(f"  [Page {page_num + 1}] No text detected")

    doc.close()
    return "\n".join(all_text)


def create_excel(results: list[dict], output_path: str):
    """
    Create a styled Excel workbook with the extraction results.
    Column A = File Name, Column B = UDIN
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "UDIN Extraction Results"

    # -- Styling --
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    data_font = Font(name="Calibri", size=11)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    success_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fail_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # -- Headers --
    headers = ["File Name", "UDIN"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # -- Data Rows --
    for row_idx, item in enumerate(results, 2):
        file_cell = ws.cell(row=row_idx, column=1, value=item["file_name"])
        udin_cell = ws.cell(row=row_idx, column=2, value=item["udin"] or "EXTRACTION FAILED")

        file_cell.font = data_font
        file_cell.alignment = data_alignment
        file_cell.border = thin_border

        udin_cell.font = data_font
        udin_cell.alignment = data_alignment
        udin_cell.border = thin_border

        if item["udin"]:
            file_cell.fill = success_fill
            udin_cell.fill = success_fill
        else:
            file_cell.fill = fail_fill
            udin_cell.fill = fail_fill

    # -- Column Widths --
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 30

    # -- Freeze Header --
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\n[OK] Excel saved to: {output_path}")


def main():
    print("=" * 70)
    print("UDIN Extractor")
    print("=" * 70)
    
    # Ask user for the folder path
    folder_path = input("Please enter the folder path containing the PDFs (leave blank for current folder): ").strip()
    
    if not folder_path:
        input_dir = Path(".").resolve()
    else:
        # Strip quotes if the user pasted a path with quotes
        folder_path = folder_path.strip('"\'')
        input_dir = Path(folder_path).resolve()

    if not input_dir.exists() or not input_dir.is_dir():
        print(f"[X] Directory does not exist: {input_dir}")
        sys.exit(1)

    output_path = input_dir / "udin_extraction_results.xlsx"
    failed_dir = input_dir / "failed extraction"

    # Find all PDF files
    pdf_files = sorted(input_dir.glob("*.pdf"))
    if not pdf_files:
        print(f"[X] No PDF files found in: {input_dir}")
        sys.exit(1)

    print(f"[DIR]   Input directory : {input_dir}")
    print(f"[PDF]   PDF files found : {len(pdf_files)}")
    print(f"[XLS]   Output Excel    : {output_path}")
    print(f"[FAIL]  Failed folder   : {failed_dir}")
    print("=" * 70)

    # Initialize RapidOCR engine
    print("\n[INIT] Initializing RapidOCR engine...")
    ocr_engine = RapidOCR()

    results = []
    failed_files = []

    for i, pdf_path in enumerate(pdf_files, 1):
        file_name = pdf_path.name
        print(f"\n[{i}/{len(pdf_files)}] Processing: {file_name}")
        print("-" * 50)

        # Run OCR
        ocr_text = ocr_pdf(str(pdf_path), ocr_engine)

        if not ocr_text.strip():
            print(f"  [WARN] No text extracted from OCR")
            results.append({"file_name": file_name, "udin": None})
            failed_files.append(pdf_path)
            continue

        # Show a snippet of OCR text for debugging
        snippet = ocr_text[:200].replace("\n", " ")
        print(f"  [TEXT] OCR Preview: {snippet}...")

        # Extract UDIN
        udin = extract_udin_from_text(ocr_text)

        if udin:
            print(f"  [OK] UDIN Found: {udin}")
            results.append({"file_name": file_name, "udin": udin})
        else:
            print(f"  [FAIL] UDIN NOT found in OCR text")
            # Dump full OCR text for debugging
            print(f"  [DEBUG] Full OCR text dump:")
            print(f"  {ocr_text}")
            results.append({"file_name": file_name, "udin": None})
            failed_files.append(pdf_path)

    # -- Create Excel --
    print("\n" + "=" * 70)
    create_excel(results, str(output_path))

    # -- Move failed files --
    if failed_files:
        failed_dir.mkdir(exist_ok=True)
        print(f"\n[MOVE] Moving {len(failed_files)} failed file(s) to: {failed_dir}")
        for f in failed_files:
            dest = failed_dir / f.name
            shutil.move(str(f), str(dest))
            print(f"  -> Moved: {f.name}")
        print(f"\n[NOTE] Failed files have been moved to '{failed_dir.name}/'")

    # -- Summary --
    success_count = sum(1 for r in results if r["udin"])
    fail_count = sum(1 for r in results if not r["udin"])
    print(f"\n{'=' * 70}")
    print(f"SUMMARY")
    print(f"   Total PDFs processed : {len(results)}")
    print(f"   UDIN extracted [OK]  : {success_count}")
    print(f"   Failed [X]           : {fail_count}")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
